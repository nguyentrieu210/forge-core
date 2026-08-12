#!/usr/bin/env python3
"""Build searchable Markdown snapshots from the canonical Alumdoor Office sources.

Run with the bundled Codex Python runtime because it already provides openpyxl
and python-docx. The generated files intentionally preserve source paths,
timestamps and SHA-256 hashes so a later reader can detect stale snapshots.
"""

from __future__ import annotations

import hashlib
import re
import unicodedata
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, Sequence
from zipfile import ZipFile

from docx import Document
from docx.table import Table
from docx.text.paragraph import Paragraph
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


REPO = Path(__file__).resolve().parents[2]
DOWNLOADS = Path.home() / "Downloads"
OUTPUT = REPO / "docs" / "source-data"
GENERATED_AT = datetime.now().astimezone().isoformat(timespec="seconds")


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def ascii_key(value: object) -> str:
    text = unicodedata.normalize("NFD", str(value or ""))
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    return re.sub(r"\s+", " ", text.replace("\n", " ")).strip().upper()


def display_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, ".12g")
    return str(value).strip()


def md_escape(value: object) -> str:
    return display_value(value).replace("\\", "\\\\").replace("|", "\\|").replace("\r", "").replace("\n", "<br>")


def unique_headers(values: Sequence[object]) -> list[str]:
    headers: list[str] = []
    seen: Counter[str] = Counter()
    for index, value in enumerate(values):
        base = display_value(value) or get_column_letter(index + 1)
        seen[base] += 1
        headers.append(base if seen[base] == 1 else f"{base} ({seen[base]})")
    return headers


def markdown_table(headers: Sequence[object], rows: Iterable[Sequence[object]]) -> list[str]:
    header_cells = [md_escape(value) for value in headers]
    lines = [
        "| " + " | ".join(header_cells) + " |",
        "| " + " | ".join("---" for _ in header_cells) + " |",
    ]
    width = len(header_cells)
    for row in rows:
        values = list(row[:width]) + [""] * max(0, width - len(row))
        lines.append("| " + " | ".join(md_escape(value) for value in values[:width]) + " |")
    return lines


def nonempty_rows(ws) -> list[tuple[int, list[object]]]:
    rows: list[tuple[int, list[object]]] = []
    for row_number, row in enumerate(ws.iter_rows(values_only=True), 1):
        values = list(row)
        while values and values[-1] in (None, ""):
            values.pop()
        if any(value not in (None, "") for value in values):
            rows.append((row_number, values))
    return rows


def source_note(path: Path) -> list[str]:
    stat = path.stat()
    return [
        f"- Nguồn: `{path.as_posix()}`",
        f"- Cập nhật nguồn: `{datetime.fromtimestamp(stat.st_mtime).astimezone().isoformat(timespec='seconds')}`",
        f"- SHA-256: `{sha256(path)}`",
        f"- Sinh Markdown: `{GENERATED_AT}`",
        "",
    ]


def find_download(*needles: str) -> Path:
    wanted = [ascii_key(value) for value in needles]
    matches = [
        path
        for path in DOWNLOADS.iterdir()
        if path.is_file() and all(needle in ascii_key(path.name) for needle in wanted)
    ]
    if not matches:
        raise FileNotFoundError(f"Không tìm thấy file Downloads chứa: {needles}")
    return max(matches, key=lambda path: path.stat().st_mtime)


def write_markdown(filename: str, lines: Sequence[str]) -> None:
    OUTPUT.mkdir(parents=True, exist_ok=True)
    content = "\n".join(lines).rstrip() + "\n"
    (OUTPUT / filename).write_text(content, encoding="utf-8", newline="\n")


def iter_doc_blocks(document: Document):
    for child in document.element.body.iterchildren():
        if isinstance(child, CT_P):
            yield Paragraph(child, document)
        elif isinstance(child, CT_Tbl):
            yield Table(child, document)


def extract_docx_media(path: Path, asset_prefix: str) -> list[str]:
    assets = OUTPUT / "assets"
    assets.mkdir(parents=True, exist_ok=True)
    with ZipFile(path) as archive:
        entries = [entry for entry in archive.infolist() if entry.filename.startswith("word/media/") and not entry.is_dir()]
        entries.sort(key=lambda entry: [int(part) if part.isdigit() else part for part in re.split(r"(\d+)", entry.filename)])
        output_names: list[str] = []
        for index, entry in enumerate(entries, 1):
            extension = Path(entry.filename).suffix.lower() or ".bin"
            output_name = f"{asset_prefix}-{index:02d}{extension}"
            (assets / output_name).write_bytes(archive.read(entry))
            output_names.append(output_name)
    return output_names


def docx_to_markdown(path: Path, title: str, asset_prefix: str) -> list[str]:
    document = Document(path)
    lines = [f"# {title}", "", *source_note(path)]
    for block in iter_doc_blocks(document):
        if isinstance(block, Paragraph):
            text = block.text.strip()
            if not text:
                continue
            style = block.style.name if block.style else ""
            heading = re.match(r"Heading\s+(\d+)", style, re.IGNORECASE)
            if heading:
                level = min(6, int(heading.group(1)) + 1)
                lines.extend([f"{'#' * level} {text}", ""])
            elif "List" in style:
                lines.append(f"- {text}")
            else:
                lines.extend([text, ""])
        else:
            rows = [[cell.text.strip() for cell in row.cells] for row in block.rows]
            if not rows:
                continue
            width = max(len(row) for row in rows)
            headers = unique_headers(rows[0] + [""] * (width - len(rows[0])))
            lines.extend(markdown_table(headers, rows[1:]))
            lines.append("")
    media = extract_docx_media(path, asset_prefix)
    if media:
        lines.extend(["## Hình/đối tượng nhúng", "", f"Tài liệu gốc có **{len(media)}** hình/đối tượng nhúng:", ""])
        for index, output_name in enumerate(media, 1):
            if Path(output_name).suffix in {".png", ".jpg", ".jpeg", ".gif", ".webp", ".svg"}:
                lines.extend([f"### Hình {index}", "", f"![Hình nhúng {index}](assets/{output_name})", ""])
            else:
                lines.append(f"- [Đối tượng nhúng {index}](assets/{output_name})")
        lines.append("")
    return lines


def material_catalog(path: Path) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rows = nonempty_rows(ws)
    headers = unique_headers(rows[0][1])
    data = [values for _, values in rows[1:]]
    groups = Counter(display_value(row[2]) or "(trống)" for row in data)
    uoms = Counter(display_value(row[3]) or "(trống)" for row in data)
    lines = ["# Danh mục sản phẩm gốc", "", *source_note(path)]
    lines.extend(["## Tóm tắt", "", f"- Số dòng sản phẩm: **{len(data)}**", "- Nhóm sản phẩm:"])
    lines.extend(f"  - {name}: {count}" for name, count in groups.most_common())
    lines.append("- Đơn vị tính:")
    lines.extend(f"  - {name}: {count}" for name, count in uoms.most_common())
    lines.extend(["", "## Dữ liệu", ""])
    lines.extend(markdown_table(headers, data))
    lines.append("")
    return lines


def small_workbook(path: Path, title: str) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    lines = [f"# {title}", "", *source_note(path)]
    for ws in wb.worksheets:
        rows = nonempty_rows(ws)
        lines.extend([f"## Sheet: {ws.title}", "", f"- Dòng có dữ liệu: {len(rows)}", ""])
        if not rows:
            continue
        width = max(len(values) for _, values in rows)
        headers = ["Dòng Excel"] + [get_column_letter(index + 1) for index in range(width)]
        table_rows = [[number, *values] for number, values in rows]
        lines.extend(markdown_table(headers, table_rows))
        lines.append("")
    return lines


def inventory_snapshot(path: Path) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    lines = ["# Tồn nhôm chuẩn hóa", "", *source_note(path)]
    output_rows: list[list[object]] = []
    colors: Counter[str] = Counter()
    conditions: Counter[str] = Counter()
    statuses: Counter[str] = Counter()
    by_item: Counter[str] = Counter()
    short_not_scrap = 0
    skipped = {"MAU", "LICH_SU", "LICH SU"}
    for ws in wb.worksheets:
        if ascii_key(ws.title) in skipped:
            continue
        header_values = [cell.value for cell in ws[9]][:12]
        if not any("NGAY NHAP" in ascii_key(value) for value in header_values):
            continue
        headers = unique_headers(header_values)
        for row_number, row in enumerate(ws.iter_rows(min_row=10, max_col=12, values_only=True), 10):
            values = list(row)
            if not any(value not in (None, "", False) for value in values):
                continue
            record = dict(zip(headers, values))
            normalized = {ascii_key(key): value for key, value in record.items()}
            color = next((value for key, value in normalized.items() if key == "MAU"), None)
            condition = next((value for key, value in normalized.items() if "TINH TRANG" in key), None)
            length = next((value for key, value in normalized.items() if key.startswith("KHO")), None)
            status = next((value for key, value in normalized.items() if "THEO DOI TON" in key), None)
            scrap = next((value for key, value in normalized.items() if "PHE" in key), None)
            if color not in (None, ""):
                colors[display_value(color).upper()] += 1
            if condition not in (None, ""):
                conditions[display_value(condition).upper()] += 1
            if status not in (None, ""):
                statuses[display_value(status).upper()] += 1
            if isinstance(length, (int, float)) and float(length) < 0.25 and ascii_key(scrap) != "PHE":
                short_not_scrap += 1
            by_item[ws.title.strip()] += 1
            output_rows.append([ws.title.strip(), row_number, *values])
    lines.extend([
        "## Tóm tắt nghiệp vụ",
        "",
        f"- Dòng tồn/lô có dữ liệu: **{len(output_rows)}**",
        f"- Dòng có chiều dài dưới 0,25 m nhưng không ghi `PHẾ`: **{short_not_scrap}**",
        "- Màu xuất hiện nhiều nhất: " + ", ".join(f"{name} ({count})" for name, count in colors.most_common(10)),
        "- Tình trạng: " + ", ".join(f"{name} ({count})" for name, count in conditions.most_common()),
        "- Trạng thái tồn: " + ", ".join(f"{name} ({count})" for name, count in statuses.most_common()),
        "",
        "## Số dòng theo mã/sheet",
        "",
    ])
    lines.extend(markdown_table(["Mã/sheet", "Số dòng"], by_item.most_common()))
    lines.extend(["", "## Dữ liệu lô/tồn", ""])
    lines.extend(markdown_table(
        ["Mã/sheet", "Dòng Excel", "Ngày nhập", "Màu/Loại", "Tình trạng/Màu", "Khổ (m)", "Số lá/cây", "Ngày nhập lại", "Theo dõi tồn", "Chọn cắt", "LM/Phế", "Tổng kg", "Nhập/Ghi chú"],
        output_rows,
    ))
    lines.append("")
    return lines


def header_index(headers: Sequence[object], *needles: str) -> int | None:
    normalized = [ascii_key(value) for value in headers]
    for needle in needles:
        wanted = ascii_key(needle)
        for index, value in enumerate(normalized):
            if wanted in value:
                return index
    return None


def value_at(row: Sequence[object], index: int | None) -> object:
    return row[index] if index is not None and index < len(row) else None


def sales_snapshot(path: Path) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    lines = ["# Lịch sử đơn hàng và xuất hàng", "", *source_note(path)]
    output_rows: list[list[object]] = []
    sheet_counts: Counter[str] = Counter()
    pattern_counts: Counter[str] = Counter()
    for ws in wb.worksheets:
        iterator = ws.iter_rows(values_only=True)
        try:
            headers = list(next(iterator))
        except StopIteration:
            continue
        order_index = header_index(headers, "ĐƠN HÀNG")
        delivery_index = header_index(headers, "PHIẾU XUẤT KHO")
        if order_index is None and delivery_index is None:
            continue
        indexes = {
            "date": header_index(headers, "NGÀY ĐẶT HÀNG", "NĂM"),
            "document": header_index(headers, "SỐ CHỨNG TỪ"),
            "customer": header_index(headers, "ĐẠI LÝ"),
            "owner": header_index(headers, "NGƯỜI PHỤ TRÁCH"),
            "type": header_index(headers, "LOẠI HÀNG", "LOẠI", "TÊN SP"),
            "item": header_index(headers, "MÃ HÀNG", "MẪ HÀNG", "MẪU HÀNG"),
            "error": header_index(headers, "LỖI"),
            "note": header_index(headers, "GHI CHÚ"),
            "delivery_date": header_index(headers, "NGÀY GIAO"),
            "status": header_index(headers, "TRẠNG THÁI", "LỆNH XUẤT KHO"),
            "warehouse": header_index(headers, "KHO"),
        }
        for row_number, row in enumerate(iterator, 2):
            order_text = display_value(value_at(row, order_index))
            delivery_text = display_value(value_at(row, delivery_index))
            if not order_text and not delivery_text:
                continue
            searchable = ascii_key(order_text or delivery_text)
            for label, pattern in {
                "có cao và rộng": r"(CPB|CAO).{0,30}(RPB|RONG)",
                "có màu": r"\b(GS|VK|XN|XF|TRANG|THO|XAM|CAFE)\b",
                "có cây": r"\bCAY\b",
                "có bộ": r"\bBO\b",
                "có cặp": r"\bCAP\b",
                "có cái": r"\bCAI\b",
            }.items():
                if re.search(pattern, searchable):
                    pattern_counts[label] += 1
            sheet_counts[ws.title] += 1
            output_rows.append([
                ws.title,
                row_number,
                value_at(row, indexes["date"]),
                value_at(row, indexes["document"]),
                value_at(row, indexes["customer"]),
                value_at(row, indexes["owner"]),
                value_at(row, indexes["type"]),
                value_at(row, indexes["item"]),
                value_at(row, indexes["error"]),
                order_text,
                delivery_text,
                value_at(row, indexes["note"]),
                value_at(row, indexes["delivery_date"]),
                value_at(row, indexes["status"]),
                value_at(row, indexes["warehouse"]),
            ])
    lines.extend([
        "## Tóm tắt",
        "",
        f"- Dòng đơn/xuất chuẩn hóa: **{len(output_rows)}**",
        "- Mẫu dữ liệu nhận diện: " + ", ".join(f"{name} ({count})" for name, count in pattern_counts.most_common()),
        "",
        "## Số dòng theo sheet",
        "",
    ])
    lines.extend(markdown_table(["Sheet", "Số dòng"], sheet_counts.items()))
    lines.extend(["", "## Dữ liệu", ""])
    lines.extend(markdown_table(
        ["Sheet", "Dòng Excel", "Ngày đặt", "Số chứng từ", "Đại lý", "Phụ trách", "Loại", "Mã/Tên hàng", "Lỗi", "Đơn hàng", "Phiếu xuất", "Ghi chú", "Ngày giao", "Trạng thái", "Kho"],
        output_rows,
    ))
    lines.append("")
    return lines


def workbook_inventory(path: Path, title: str, selected_sheets: set[str] | None = None) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    lines = [f"# {title}", "", *source_note(path), "## Danh sách sheet", ""]
    inventory: list[list[object]] = []
    extracted: list[tuple[str, list[tuple[int, list[object]]]]] = []
    selected_normalized = {ascii_key(name) for name in selected_sheets or set()}
    for ws in wb.worksheets:
        rows = nonempty_rows(ws)
        width = max((len(values) for _, values in rows), default=0)
        inventory.append([ws.title, len(rows), width])
        if ascii_key(ws.title) in selected_normalized:
            extracted.append((ws.title, rows))
    lines.extend(markdown_table(["Sheet", "Dòng có dữ liệu", "Số cột tối đa"], inventory))
    for sheet_name, rows in extracted:
        lines.extend(["", f"## Trích đầy đủ sheet: {sheet_name}", ""])
        width = max((len(values) for _, values in rows), default=0)
        lines.extend(markdown_table(
            ["Dòng Excel"] + [get_column_letter(index + 1) for index in range(width)],
            [[number, *values] for number, values in rows],
        ))
    lines.append("")
    return lines


def source_manifest(sources: Sequence[tuple[str, Path, str]]) -> list[str]:
    lines = [
        "# Dữ liệu nguồn Alumdoor",
        "",
        "Thư mục này là bản Markdown tra nhanh sinh từ Word/Excel gốc. Khi số SHA-256 của file gốc thay đổi, chạy lại `server/scripts/build-alumdoor-source-md.py`.",
        "",
        "## Thứ tự ưu tiên khi có mâu thuẫn",
        "",
        "1. Excel/Word gốc có ngày sửa mới nhất.",
        "2. Dữ liệu giao dịch thực tế trong tồn nhôm và đơn/xuất.",
        "3. File đối chiếu hoặc dữ liệu đã import trong repo.",
        "4. Dữ liệu demo/fixture chỉ dùng khi nguồn thật không có.",
        "",
        "## Chỉ mục",
        "",
    ]
    rows: list[list[object]] = []
    for label, path, output_name in sources:
        stat = path.stat()
        rows.append([
            label,
            path.as_posix(),
            stat.st_size,
            datetime.fromtimestamp(stat.st_mtime).astimezone().isoformat(timespec="seconds"),
            sha256(path),
            f"[{output_name}]({output_name})" if output_name else "Chỉ lưu cấu trúc/metadata",
        ])
    lines.extend(markdown_table(["Nguồn", "Đường dẫn", "Byte", "Cập nhật", "SHA-256", "Markdown"], rows))
    lines.extend([
        "",
        "## Quy tắc bảo mật",
        "",
        "- Danh mục vật tư, quy cách, quy trình, tồn nhôm và lịch sử đơn/xuất được lưu để tìm kiếm nhanh.",
        "- Công nợ, tài khoản ngân hàng và danh sách khách hàng chỉ lưu metadata/cấu trúc, không nhân bản toàn bộ dữ liệu nhạy cảm sang Markdown.",
        "- File `SỔ NỢ (1).xlsx` thực chất là định dạng Excel BIFF/OLE cũ dù mang đuôi `.xlsx`; giữ trong chỉ mục nhưng không trích bằng openpyxl.",
        "",
    ])
    return lines


def main() -> None:
    process_doc = find_download("25.7 QUY TRÌNH", "(1)")
    order_doc = find_download("ĐƠN ĐẶT HÀNG")
    catalog = find_download("DANH MỤC SẢN PHẨM")
    rules = find_download("QUY CÁCH")
    debt = find_download("SỔ NỢ", "(1)")
    purchase_order = find_download("ĐƠN MUA HÀNG-20260730-1550")
    legacy = find_download("MS LIÊN BS")
    reconciliation = find_download("ALUMDOOR", "DOI CHIEU THANH PHAM")
    item_export = find_download("HÀNG HOÁ", "VẬT TƯ-20260728-2018")
    customer_export = REPO / "data" / "customer-export.xlsx"
    sales = REPO / "data" / "don-hang-xuat-hang.xlsx"
    sau_hong = REPO / "data" / "sau-hong.xlsx"
    stock = REPO / "data" / "ton-nhom.xlsx"

    write_markdown("quy-trinh.md", docx_to_markdown(process_doc, "Quy trình Alumdoor — nguồn Word", "quy-trinh"))
    write_markdown("don-dat-hang-template.md", docx_to_markdown(order_doc, "Đơn đặt hàng — nguồn Word", "don-dat-hang"))
    write_markdown("danh-muc-san-pham.md", material_catalog(catalog))
    write_markdown("quy-cach-va-mau.md", small_workbook(rules, "Quy cách sản xuất, mua và bán; danh mục màu"))
    write_markdown("ton-nhom.md", inventory_snapshot(stock))
    write_markdown("don-hang-xuat-hang.md", sales_snapshot(sales))
    write_markdown("doi-chieu-thanh-pham.md", small_workbook(reconciliation, "Đối chiếu thành phẩm"))
    write_markdown("hang-hoa-vat-tu-import.md", small_workbook(item_export, "Hàng hóa / Vật tư đã xuất từ hệ thống"))
    write_markdown("don-mua-hang.md", small_workbook(purchase_order, "Đơn mua hàng mẫu"))
    write_markdown("sau-hong.md", small_workbook(sau_hong, "Báo giá và đơn của CTY Sáu Hồng"))
    write_markdown(
        "du-lieu-ke-toan-cu.md",
        workbook_inventory(legacy, "Dữ liệu kế toán cũ — phần nghiệp vụ được phép trích", {"GHI CHÚ", "Trang tính29", "theo dõi chi tiết vật tư"}),
    )
    sources = [
        ("Quy trình sản xuất", process_doc, "quy-trinh.md"),
        ("Mẫu đơn đặt hàng", order_doc, "don-dat-hang-template.md"),
        ("Danh mục sản phẩm", catalog, "danh-muc-san-pham.md"),
        ("Quy cách và màu", rules, "quy-cach-va-mau.md"),
        ("Tồn nhôm", stock, "ton-nhom.md"),
        ("Đơn hàng và xuất hàng", sales, "don-hang-xuat-hang.md"),
        ("Đối chiếu thành phẩm", reconciliation, "doi-chieu-thanh-pham.md"),
        ("Hàng hóa / Vật tư export", item_export, "hang-hoa-vat-tu-import.md"),
        ("Đơn mua hàng", purchase_order, "don-mua-hang.md"),
        ("Báo giá CTY Sáu Hồng", sau_hong, "sau-hong.md"),
        ("Dữ liệu kế toán cũ", legacy, "du-lieu-ke-toan-cu.md"),
        ("Sổ nợ", debt, ""),
        ("Khách hàng export", customer_export, ""),
    ]
    write_markdown("README.md", source_manifest(sources))
    print(f"ALUMDOOR_SOURCE_MD_PASS files={len(list(OUTPUT.glob('*.md')))} output={OUTPUT}")


if __name__ == "__main__":
    main()
